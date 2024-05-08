
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

wb = xl.Workbook()

ws = wb.active
ws.title = 'Movies Summary'

headers = ['No.', 'Movie Title', 'Release Date', 'Total Gross', 'Theaters', 'Average per Theater']
fontobject = Font(name='Times New Roman', size=12, bold=True)
ws.append(headers)
for cell in ws[1]:
    cell.font = fontobject

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws.column_dimensions[col].width = 20

movies_data = [
    (1, "Dune: Part Two", "Mar 1", 277239255, 4074),
    (2, "Kung Fu Panda 4", "Mar 8", 180792675, 4067),
    (3, "Godzilla x Kong: The New Empire", "Mar 29", 173230709, 3948),
    (4, "Ghostbusters: Frozen Empire", "Mar 22", 103596334, 4345),
    (5, "Bob Marley: One Love", "Feb 14", 96870413, 3597)
]

for movie in movies_data:
    ws.append(movie)
    last_row = ws.max_row
    total_gross_cell = 'D' + str(last_row)
    theaters_cell = 'E' + str(last_row)
    average_formula = f"={total_gross_cell}/{theaters_cell}"
    ws['F' + str(last_row)] = average_formula

    ws[total_gross_cell].number_format = '#,##0'
    ws['F' + str(last_row)].number_format = '#,##0'


wb.save('MoviesList.xlsx')

